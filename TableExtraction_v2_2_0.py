# -*- coding: utf-8 -*-
#import os
import cv2
import numpy as np
import os
from geometry import Rect
from geometry import Point
from operator import itemgetter
import matplotlib.pyplot as plt
import openpyxl 
import LineDetection as Imaging
import whiteSpace
isDebugImg = True


class Table():  
    tablecoordinates = []
    tablerows = []
    tablecols = []
    hrPoints = []
    vrPoints = []
    tableHeaders = []
    def __init__(self):
        self.tablerows = []
        self.tablecols = []
        self.tablecoordinates = []
        self.hrPoints = []
        self.vrPoints = []
        self.tableHeaders = []
    def settablecoordinates(self,eachtable):
        self.tablecoordinates.append(eachtable[0])
        self.tablecoordinates.append(eachtable[1])
        self.tablecoordinates.append(eachtable[2])
        self.tablecoordinates.append(eachtable[3])
        
    def settablerows(self, eachtablerow):
        self.tablerows.append(eachtablerow)
    
    def settablecols(self,eachtablecol):
        self.tablecols.append(eachtablecol)
        
    def sethrPoints(self,hrpoint):
        self.hrPoints.append(hrpoint)
        
    def setvrPoints(self,vrpoint):
        self.vrPoints.append(vrpoint)

    def settableHeaders(self,firstRowData):
        for eachrow in firstRowData:
            self.tableHeaders.append(eachrow)

def funf(ca, cb):    
    if ca[1] <= (cb[1] + cb[3]) and (ca[1] + ca[3]) >= cb[1]:        
        return True
    return False

def morphCloseBlobs(binary):
    closingblobs = []
    contours, hierarchy = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
    for contour in contours:
            x, y, w, h = cv2.boundingRect(contour)            
            closingblobs.append([x, y, w, h])
    return closingblobs
    
def Vlineblobfillfun(colblobimg, colblobs, tablecoordinates): 
    rowx = tablecoordinates[0] 
    rowy = tablecoordinates[1]
    rowW = tablecoordinates[2]
    rowH = tablecoordinates[3]
    r1 = Rect(rowx, rowy ,rowW ,rowH)
    vblobimg = colblobimg.copy()
    
    for ca in colblobs:
        vx = ca[0]; vy = ca[1];
        vw = ca[2] - ca[0]
        vh = ca[3] - ca[1]
        r2 = Rect(vx,vy,vw,vh) 
        midpointx = int(vw/2); midpointy = int(vh/2)
        pointr2 = Point(vx  + midpointx , vy + midpointy)                    
        res =  r1.overlaps_with(r2)                         
        if res:                       
            respoint = r1.is_point_inside_rect(pointr2)
            if respoint: 
                cv2.rectangle(vblobimg, (ca[0], ca[1]), (ca[2], ca[3]), 255, cv2.FILLED)           
    return vblobimg

def lineblobfillfun(thresh, ca, cb):
    if funf(ca,cb):
        x = ca[0] + ca[2]
        y = ca[1]
        w = cb[0] - x
        h = ca[3]
        cv2.rectangle(thresh, (x, y), (x+w, y+h), 255, cv2.FILLED)
    return thresh

def borderNoiseRemoval(binary, border):
    h, w = binary.shape
    cv2.rectangle(binary, (0,0), (w, border), 255, cv2.FILLED)
    cv2.rectangle(binary, (0, 0), (border, h), 255, cv2.FILLED)
    cv2.rectangle(binary, (0, h-border), (w, h), 255, cv2.FILLED)
    cv2.rectangle(binary, (w-border, 0), (w, h), 255, cv2.FILLED)           
    return binary

def second_Hump(hist):
    length = len(hist)
    flag = 0
    first_Hump = max(hist)
    if hist[0] == first_Hump:
        flag = 1
    peak = []
    for i in range(1,length-1):
        if (hist[i] > hist[i-1] and hist[i] > hist[i+1]):
            peak.append(hist[i])
#    print(len(peak))
    if(len(peak) > 2):
        if flag:
            second_hump = max(peak)
        else:
            peak.sort()
            second_hump = peak[len(peak)-2]  
         
        new_hist = [float(i) for i in hist]
        hump_pos = new_hist.index(second_hump)
        try:
            while new_hist[hump_pos] > 0:
    	        hump_pos += 1
            hump_pos -= 1
        except:
            pass
        
        #print("inside second hum h1:{}".format(new_hist))
        return second_hump, hump_pos
    else:
        return 10,0

def prominentHump(h3):
    new_hist = list(h3)
    prom_Hump = max(new_hist)
    hump_Position = new_hist.index(prom_Hump)
    hist_len = len(new_hist) 
    if hump_Position == hist_len-1:
        return hump_Position  
    while new_hist[hump_Position] > 0 and hump_Position < (len(new_hist)-1):
        hump_Position += 1
    return hump_Position-1

#def getGmax(linegap):
#    tableRowGap = list(filter(lambda x: x < 60, linegap))
#    h, bins, patches = plt.hist(tableRowGap)
#    ind = prominentHump(h)
#    gmax = round((bins[ind] + bins[ind+1])/2)+3
#    return gmax
    
def getGmax(tablelines):
    linegap = []
    length = len(tablelines)
    for i in range(0, length-1):
        x, y, w, h = stats2rect(tablelines[i])
        bottom_TELa = tablelines[i][3]
        top_TELa1 = tablelines[i+1][1]
        gap = round((top_TELa1 - bottom_TELa))
        linegap.append(gap)
    tableRowGap = list(filter(lambda x: x < 70, linegap))
    h, bins, patches = plt.hist(tableRowGap)
    ind = prominentHump(h)
    gmax = round((bins[ind] + bins[ind+1])/2)+3
    return gmax


def tableCols(tablerect, tablerowrect):
    FinalTables = []
    tables = []
    tablec = 0
    for eachblob in tablerect:
            doc_table = Table()
            doc_table.settablecoordinates(eachblob)      
            tablec = tablec + 1
            r1 = Rect(eachblob[0], eachblob[1], (eachblob[2]), (eachblob[3]))                
            for i, hrblob in enumerate(tablerowrect):                 
                r2 = Rect(hrblob[0],hrblob[1],(hrblob[2] - hrblob[0]),(hrblob[3] - hrblob[1])) 
                midpointx = int((hrblob[2] - hrblob[0])/2); midpointy = int((hrblob[3] - hrblob[1])/2)
                pointr2 = Point(hrblob[0] + midpointx , hrblob[1] + midpointy)                    
                res =  r1.overlaps_with(r2)                 
                if res:   
                    respoint = r1.is_point_inside_rect(pointr2)
                    if respoint:
                        #cv2.rectangle(tableimg, (hrblob[0]+5, hrblob[1]), (hrblob[2] - 5, hrblob[3]), (255,191,0 ), 2)
                        hrblob[2] = eachblob[0] + eachblob[2]
                        tables.append([hrblob, tablec]) 
                        doc_table.settablerows(hrblob)
            FinalTables.append(doc_table)
            del doc_table
    return tables , FinalTables 

def lineRemoval(binary):
    blurred = cv2.GaussianBlur(binary, (5, 5), 11)    
    thresh = cv2.adaptiveThreshold(~blurred, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 25, -2)
    scale = 10 
    scale2 = 35 
    horizontal = thresh.copy()
    vertical = thresh.copy()    
    height, width = horizontal.shape[:2]
    horizontalsize = round(width / scale)    
    horizontalStructure = cv2.getStructuringElement(cv2.MORPH_RECT, (horizontalsize,1))
    horizontal = cv2.erode(horizontal, horizontalStructure, iterations=1)
    horizontal = cv2.dilate(horizontal, horizontalStructure, iterations=1)
    verticalsize = round(height / scale2)
    verticalStructure = cv2.getStructuringElement(cv2.MORPH_RECT, (1, verticalsize))
    vertical = cv2.erode(vertical, verticalStructure, iterations=1)
    vertical = cv2.dilate(vertical, verticalStructure, iterations=1)
    h_contours,hh = cv2.findContours(horizontal,cv2.RETR_TREE,cv2.CHAIN_APPROX_NONE)
    for contour in h_contours:
        x, y, w, h = cv2.boundingRect(contour)            
        cv2.rectangle(binary, (x, y), (x+w, y+h), 255, cv2.FILLED)
    
    v_contours,hv = cv2.findContours(vertical,cv2.RETR_TREE,cv2.CHAIN_APPROX_NONE)
    for contour in v_contours:
        x, y, w, h = cv2.boundingRect(contour)            
        cv2.rectangle(binary, (x, y), (x+w, y+h), 255, cv2.FILLED)
    
    kernel = np.ones((3,3), np.uint8)
#    linesRemovedImg = cv2.erode(binary, kernel) # to reduce time
    linesRemovedImg = ~(cv2.dilate(~binary, kernel))
    return linesRemovedImg

def stats2rect(ca):
    x = ca[0]
    y = ca[1]
    w = ca[2]
    h = ca[3]
    return x, y, w, h

def horizontalProjection(word_Blobs_Image, word_Blobs):
    imgw, imgh = word_Blobs_Image.shape
    for i,ca in enumerate(word_Blobs): 
        if (ca[2] > 5 and ca[3] > 10 and ca[3] < (imgh*0.4)):   #check        
#           cv2.rectangle(inputImg, (ca[0], ca[1]), (imgw, ca[1]+ca[3]), (255, 0, 0), 1)
           bloba = [ca[0], ca[1], ca[2], ca[3]]
           r1 = Rect(ca[0], ca[1], imgw,ca[3])
           for j, cb in enumerate(word_Blobs):
               if i!=j:
                   blobb = [cb[0], cb[1], cb[2], cb[3]]
                   r2 = Rect(cb[0], cb[1], imgw,cb[3])                
                   res = r1.overlaps_with(r2)
                   if res: 
#                       cv2.rectangle(srcB, (cb[0], cb[1]), (imgw, cb[1]+cb[3]), (255, 0, 0), 1)
                       word_Blobs_Image = lineblobfillfun(word_Blobs_Image, bloba, blobb)
    
    #cv2.imwrite('hr_prj.png',word_Blobs_Image)
    return word_Blobs_Image

def verticalProjection(tableimg, tablecoordinates, tablerowrect, allImageBlobs, blank_image):
    vblobs = []
    tableH = tablecoordinates[1] + tablecoordinates[3] 
    vvblobs = []
    for i,eachblob in enumerate(tablerowrect): #for eachblob in tablerowrect:
        rowx = eachblob[0]
        rowy = eachblob[1]
        rowW = eachblob[2] - eachblob[0]
        rowH = eachblob[3] - eachblob[1]
        r1 = Rect(rowx, rowy, rowW ,rowH)
        cv2.rectangle(tableimg, (rowx, rowy), (rowx + rowW,rowy+rowH), (255,0,0 ), 2)
        for i, vrblob in enumerate(allImageBlobs):
            vx = vrblob[0]
            vy = vrblob[1]
            vw = vrblob[2]
            vh = vrblob[3]
            #cv2.rectangle(tableimg, (vx, vy), (vx+vw, vy+vh), (0,255,0 ), 2)
            r2 = Rect(vx, vy, vw, vh)
            midpointx = int(vw/2)
            midpointy = int(vh/2)
            pointr2 = Point(vx+midpointx ,vy+midpointy)
            res =  r1.overlaps_with(r2)
            if res:
                respoint = r1.is_point_inside_rect(pointr2)
                if respoint:
                    vvblobs.append(vrblob)
                    vx = vrblob[0]
                    vy = vrblob[1]
                    r = vrblob[0] + vrblob[2]
                    h = tableH
                    cv2.rectangle(tableimg, (vx, vy), (r, h), (0,0,255 ), 2)
                    cv2.rectangle(blank_image, (vx, vy), (r,h), 255, cv2.FILLED)   
                    vblobs.append([vx, vy, r, h])
#    cv2.imwrite('tableimg.png',tableimg)
#    cv2.imwrite('blankimg.png', blank_image)
#    print(vblobs)
    return vblobs, tableimg, blank_image

def preprocess(inputSrcImg):
    # smooth the image with alternative closing and opening with an enlarging kernel
#    morph = im.copy()
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 1))
    inputSrcImg = cv2.morphologyEx(inputSrcImg, cv2.MORPH_CLOSE, kernel)
    inputSrcImg = cv2.morphologyEx(inputSrcImg, cv2.MORPH_OPEN, kernel)
    preProcessedImg = np.split(np.asarray(inputSrcImg), 3, axis=2)    
    channel_height, channel_width, _ = preProcessedImg[0].shape
    
    # apply Otsu threshold to each channel
    for i in range(0, 3):
        _, preProcessedImg[i] = cv2.threshold(preProcessedImg[i], 0, 255, cv2.THRESH_OTSU | cv2.THRESH_BINARY)
        preProcessedImg[i] = np.reshape(preProcessedImg[i], newshape=(channel_height, channel_width, 1))
    
    # merge the channels
    preProcessedImg = np.concatenate((preProcessedImg[0], preProcessedImg[1], preProcessedImg[2]), axis=2)
    return preProcessedImg

def connectedComp_Analysis(linesRemovedImg):
    '''connected component analysis'''
    connectivity = 8  # You need to choose 4 or 8 for connectivity type
    stats = cv2.connectedComponentsWithStats(~linesRemovedImg, connectivity, cv2.CV_32S)[2]
    cc_stats = list(stats)
    return cc_stats

def distancefunD(ca, cb):
    d = cb[0] - (ca[0]+ca[2])
    return d       

def getWordsGap(cc_stats, src):
    imgh,imgw = src.shape[:2]
    blobsDrawn = np.zeros((imgh,imgw), np.uint8)     
        
    compB = []
    compA = []
    alldcomp = []
    hrsize = 0
    for i, ca in enumerate(cc_stats):        
        x, y, w, h = stats2rect(ca) 
        if w > 2 and h < 80:#and h>15:
            lcx = []       
            cv2.rectangle(src, (x, y), (x+w, y+h), (255,0,0), 2)
            cv2.rectangle(blobsDrawn, (x, y), (x+w, y+h), 255, cv2.FILLED) 
            for j, cb in enumerate(cc_stats):
                    if i != j:
                        funfres = funf(ca, cb)
                        if funfres:
                            d = distancefunD(ca, cb)
                            if funfres and cb[0] > (ca[0]+ca[2]):
                                 lcx.append([d, j])                                                       
            if len(lcx) > 1:
                idx = min(lcx, key=itemgetter(0))                
                alldcomp.append(idx[0])
                compB.append(cc_stats[idx[1]])
                compA.append(ca)
    
#    print("len of alldcomp:{}".format((alldcomp))) 
#    print("len of alldcomp before:{}".format(len(alldcomp)))
    alldcomp = list(filter(lambda x: x < 25, alldcomp))
#    print(alldcomp)
    num_bins = max(alldcomp)
    
    
    plt.title("WordBlob Histogram")
    plt.xlabel("Word gap")
    plt.ylabel("Frequency")
    h1, bins, patches = plt.hist(alldcomp, bins = num_bins, facecolor='green', alpha=0.5)
    
    secondHump, hrsize= second_Hump(h1)
    hrsize = bins[int(hrsize)]
    if hrsize < 10:
        hrsize  = 10
    
#    print("len of alldcomp after:{}".format(len(alldcomp)))
#    print("Num_bins:{} , h1:{}, hrsize:{}, bins:{}".format(num_bins,len(h1),hrsize,len(bins)))
    
#    exit()
    return hrsize, secondHump, src,blobsDrawn

def textLines(hrsize, secondHump, linesRemovedImg, outPath):
#    if hrsize > 30:
#        hrsize = 10
#        secondHump = 10
    
    imgh, imgw = linesRemovedImg.shape[:2]
    head, tail = os.path.split(outPath)
    filename = os.path.splitext(tail)[0] #tail[:-4]
    'Line removal'
    withoutLines = (~linesRemovedImg).copy()
#    linesBlobs = morphCloseBlobs(withoutLines)
#    for ca in linesBlobs:
#        if ca[3] < 10 :
#            cv2.rectangle(withoutLines, (ca[0], ca[1]), (ca[0]+ca[2], ca[1]+ca[3]), 0, cv2.FILLED)        
#    
#    if isDebugImg:
#            save_stats = head +"\\"+ filename + "_linesRemoved.png"
#            cv2.imwrite(save_stats, withoutLines)
            
    if hrsize > 0:
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (int(hrsize), 1))
        morph = cv2.morphologyEx(withoutLines, cv2.MORPH_CLOSE, kernel, iterations = 1)
        if True:
            save_stats = head +"\\"+ filename + "_morph.png"
            cv2.imwrite(save_stats, morph)
#        if isDebugImg:
#            save_stats = head +"\\"+ filename + "_morph_SA.png"
            
#            cv2.imwrite(save_stats, morph)
    closingblobs = morphCloseBlobs(morph)
    closingblobs.sort(key=lambda x: x)     
    word_Blob_Image = np.zeros((imgh,imgw), np.uint8)
#    secondhumpthresh = word_Blob_Image.copy()
#    colblobs = word_Blob_Image.copy()
    for ca in closingblobs:
        if (ca[2] > 5 and ca[3] >= 10 and ca[3] < (imgh * 0.4) and ca[2] < (imgw * 0.5)):
            cv2.rectangle(word_Blob_Image, (ca[0], ca[1]), (ca[0]+ca[2], ca[1]+ca[3]), 255, cv2.FILLED)
        
        
    
    wordBlobs = morphCloseBlobs(word_Blob_Image)
    wordBlobs.sort(key=lambda x: x) 
    if isDebugImg:
        save_stats = head +"\\"+ filename + "_wordsimg.png"
        cv2.imwrite(save_stats, word_Blob_Image)
#    '''second hump section'''
#    if hrsize > 0:
#        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (int(secondHump), 1))
#        #morphsecondHump = cv2.morphologyEx(~linesRemovedImg, cv2.MORPH_CLOSE, kernel,  iterations = 2)
#        morphsecondHump = cv2.morphologyEx(withoutLines, cv2.MORPH_CLOSE, kernel,  iterations = 2)
#        if isDebugImg:
#            save_stats = head +"\\"+ filename + "_secondHump.png"
#            cv2.imwrite(save_stats, morphsecondHump)
#    secondHumpblobs = morphCloseBlobs(morphsecondHump)
#    for ca in secondHumpblobs:
#        if (ca[2] > 5 and ca[3] >= 10 and ca[3] < (imgh * 0.4)):
#            cv2.rectangle(secondhumpthresh, (ca[0], ca[1]), (ca[0]+ca[2], ca[1]+ca[3]), 255, cv2.FILLED)        
#                    
#    secondHumpwordblobs = morphCloseBlobs(secondhumpthresh)
#    secondHumpwordblobs.sort(key=lambda x: x)
#    #    linesBlobs = morphCloseBlobs(withoutLines)
##    for ca in secondHumpwordblobs:
##            cv2.rectangle(colblobs, (ca[0], ca[1]), (ca[0]+ca[2], ca[1]+ca[3]), 255, cv2.FILLED)
#    if isDebugImg:
#        save_stats = head +"\\"+ filename + "_morphsecondHump.png"
#        cv2.imwrite(save_stats, secondhumpthresh)
#        save_stats = head +"\\"+ filename + "_colblob.png"
#        cv2.imwrite(save_stats, colblobs)
    return word_Blob_Image , wordBlobs #, closingblobs #, secondHumpwordblobs #--Check col blobs  closingblobs

def candidateTableRows(textlineinput, wordblobs):
#    test = cv2.imread("C:\ICDAR\Code\TRACKB2\Quicktest\cTDaR_t10022.jpg")
    textlines = morphCloseBlobs(textlineinput)
    textlines.sort(key=lambda x: x)
    tablerows = []
    for i, ca in enumerate(textlines):
        rowcount = 0 
        r1 = Rect(ca[0], ca[1], ca[2], ca[3])
        for j, cb in enumerate(wordblobs):
               r2 = Rect(cb[0], cb[1], cb[2], cb[3])
               res =  r1.overlaps_with(r2)
               if res:
                   rowcount = rowcount + 1
                   #cv2.rectangle(test, (cb[0], cb[1]), (cb[0]+cb[2], cb[1]+cb[3]), (0, 255, 0), 2)
        tablerows.append([ca,rowcount])
    tablerows.sort(key=lambda x: x[1])
#    cv2.imwrite("rows.jpg",test)
    return tablerows

def findTableRows(tablerows, srclineblob):
    tablelines = []
    singleblobs = []
    for i,ca in enumerate(tablerows):
        if ca[1] > 1:  #number of blobs in a line          
            x = ca[0][0]
            y = ca[0][1]
            r = ca[0][0] + ca[0][2]
            b = ca[0][1] + ca[0][3] 
            tablelines.append([x,y,r,b])
            cv2.rectangle(srclineblob, (x, y), (r, b), (255, 0, 0), 2)
        else:
            if ca[0][3] > 5:
                x = ca[0][0]
                y = ca[0][1]
                r = ca[0][0] + ca[0][2]
                b = ca[0][1] + ca[0][3]
#                tablelines.append([x,y,r,b])
                singleblobs.append([x,y,r,b])
				
                cv2.rectangle(srclineblob, (x, y), (r, b), (0, 0, 255), 2)
    tablelines.sort(key=lambda x: x[1])
    singleblobs.sort(key=lambda x: x[1])
    return tablelines, srclineblob, singleblobs
def rowCorrection(tablelines, singleblobs, gmax):
#    print('singleblobs', singleblobs)
#    print('tablelines', tablelines)
    sigleBlbLen = len(singleblobs)-1
    i = 0
    while i < len(tablelines):
        row_b = tablelines[i][3]
        j = 0
        while j < len(singleblobs):
                singleblob_y = singleblobs[j][1]
                vertical_dist = singleblob_y - row_b 
                if vertical_dist > 0 and vertical_dist < gmax:
                    #print('appending', singleblobs[j])
                    tablelines.append(singleblobs[j])
#                    print('tablelines',tablelines)
                    del singleblobs[j]
#                    print('singleblobs', singleblobs)
                    j = sigleBlbLen
                j += 1
        tablelines.sort(key=lambda x: x[1])
        i += 1
    tablelines.sort(key=lambda x: x[1])
#    print('tablelines', len(tablelines))
    return tablelines

def findTableCols(FinalTables, table, secondHumpwordblobs, cols_image, blank_image):
    tablecount = 0
    for i, eachtable in enumerate(FinalTables):
#         tableT = eachtable.tablecoordinates[1]
         if len(eachtable.tablerows) >= 1:
             vblobs, tablevr,blank_image =  verticalProjection(table,eachtable.tablecoordinates, eachtable.tablerows,secondHumpwordblobs,blank_image) #,closingblobs)
             tablecount = tablecount + 1
             x = eachtable.tablecoordinates[0]; y = eachtable.tablecoordinates[1];
             w = eachtable.tablecoordinates[2]; h = eachtable.tablecoordinates[3]
             cv2.rectangle(table, (x, y), (x+w, y+h), (0,255,0  ),3)
             
             for i,t in enumerate(eachtable.tablerows):
                 cv2.rectangle(table, (t[0], t[1]), (t[2],t[3]), (255,191,0), 1)
             
             cols_image1 = Vlineblobfillfun(cols_image,vblobs,eachtable.tablecoordinates)
             contours = morphCloseBlobs(cols_image1)
             contours.sort(key=lambda x: x)
             #print("sort contours:{0}".format(contours))
             for contour in contours:                   
                    x = contour[0]
                    y = y
                    w = contour[2]
                    h = h
                    cv2.rectangle(table, (x, y), ((x+w), y+h), (255,0,0),3)
                    eachtable.tablecols.append([x, y, x+w, y+h])
    return blank_image, table

def writeToExcel_ColFixed(excelsavefilepath,Tables):
    excelSavePath = excelsavefilepath
#    print("Table parser: ",Tables)
    try:
        wb = openpyxl.Workbook() 
        sheetname = "Table"    
        ws0 = wb.worksheets[0]
        ws0.title = 'Table-' + str(0) #'Table0'
        for i in range(0,len(Tables[0])):
            nco = Tables[0][i]
            for j in range(0,len(nco)):
                c1 = ws0.cell(row = i+1, column = j+1)
                c1.value = str(Tables[0][i][j])
        wb.save(excelSavePath)
        
        for t, rowCells in enumerate(Tables):
                s = wb.create_sheet(index = 1, title = sheetname) 
                ignoresheet = s.title            
                rows = len(rowCells)+1; cols = len(rowCells[0])+1 
#                print(rows, cols)
                if  ignoresheet !=  sheetname or 'Sheet' != ignoresheet:
                    s.title = 'Table-' + str(t)
                    for i in range(0,len(rowCells)):
                        nco = rowCells[i]
                        for j in range(0,len(nco)):
                            c1 = s.cell(row = i+1, column = j+1)
                            c1.value = str(rowCells[i][j])
        wb.remove_sheet(wb["Table-01"])
        excelSavePath = excelsavefilepath
        wb.save(excelSavePath)
        print("Workbook saved successfully...") 
    except IOError:
        print("Failed to save xls.")
    finally:
        excelSavePath = excelsavefilepath
        wb.save(excelSavePath)

def writeToExcel(excelsavefilepath,Tables):
    print("Write to excel is processing...")
    wb = openpyxl.Workbook() 
    sheetname = "Table"    
    ws0 = wb.worksheets[0]
    ws0.title = 'Table-' + str(0) #'Table0'
    r = len(Tables[0])
    c = len(Tables[0][0])
    rows = r+1; cols = c+1    
    for i in range(1,rows):
        for j in range(1,cols):
            c1 = ws0.cell(row = i, column = j)  
            c1.value = str(Tables[0][i-1][j-1])
    
    excelSavePath = excelsavefilepath
    wb.save(excelSavePath)
    for i, rowCells in enumerate(Tables):
            s = wb.create_sheet(index = 1, title = sheetname) 
            ignoresheet = s.title            
            rows = len(rowCells)+1; cols = len(rowCells[0])+1            
            if  ignoresheet !=  sheetname or 'Sheet' != ignoresheet:
                s.title = 'Table-' + str(i)
                for i in range(1,rows):
                    for j in range(1,cols):
                        c1 = s.cell(row = i, column = j)           
                        c1.value =str(rowCells[i-1][j-1])
    wb.remove_sheet(wb["Table-01"])   
    excelSavePath = excelsavefilepath
    wb.save(excelSavePath)

                        
def createCellBlobs(FinalTables, cellblobs_image, finalTableImg, save_stats, resultsSavePath):
#hr cell blobs
    isCellBlobs = False
    vcellblobs_image = cellblobs_image.copy()
    tablecells = finalTableImg.copy()
    tableOCRImg = finalTableImg.copy()
    ImageTables = []
    for i, eachtable in enumerate(FinalTables):
        if len(eachtable.tablerows) > 1 and len(eachtable.tablecols) > 1:
            doc_table = Table()
            doc_table.settablecoordinates(eachtable.tablecoordinates)             
            tX = eachtable.tablecoordinates[0]; tY = eachtable.tablecoordinates[1]; 
            tW = eachtable.tablecoordinates[2]; tH = eachtable.tablecoordinates[3];
            cv2.rectangle(finalTableImg, (tX, tY), (tW + tX, tH+tY), (0,0,255), 2)

            doc_table.setvrPoints(tY)

            for i in range(0, len(eachtable.tablerows)-1):
                bottom_TELa = eachtable.tablerows[i][3]
                top_TELa1 = eachtable.tablerows[i+1][1] 
                gap = 0                               
                gap = round((top_TELa1 - bottom_TELa)/2)
                hr = gap + eachtable.tablerows[i][3] 
                cv2.rectangle(finalTableImg, (tX, hr), (tX+tW, hr), (255,191,0),2)
                doc_table.settablerows([tX, hr,(tX+tW), hr])                
                doc_table.setvrPoints(hr)           
            doc_table.setvrPoints(tY+tH)
            doc_table.sethrPoints(tX)

            for i in range(0, len(eachtable.tablecols)-1):
                right_TELa = eachtable.tablecols[i][2]
                left_TELa1 = eachtable.tablecols[i+1][0] 
                gap = 0                               
                gap = round((left_TELa1 - right_TELa)/2)
                vx = eachtable.tablecols[i][2] + gap
                cv2.rectangle(vcellblobs_image, (vx, tY), (vx, tY+tH), (255,255,255),2)
                cv2.rectangle(finalTableImg, (vx, tY), (vx, tY+tH), (255,0,0),2)
                doc_table.settablecols([vx, tY, vx, (tY+tH)])                
                doc_table.sethrPoints(vx)
            doc_table.sethrPoints(tX+tW)
            ImageTables.append(doc_table) 
            print()
            del doc_table   
    totalTables = len(ImageTables) 
    print("Total tables:{0}".format(totalTables))
    
    Tables = []
    for i, eachtable in enumerate(ImageTables):
        if i == i:
            vrLen = len(eachtable.vrPoints)
            hrLen = len(eachtable.hrPoints)
            rowCells = []
            for i,vr in enumerate(eachtable.vrPoints):
                if i != vrLen-1:
                    cols = []                
                    for j,hr in enumerate(eachtable.hrPoints):
                            if i < vrLen-1 and j < hrLen-1:
                                    cell = []
                                    cell = [eachtable.hrPoints[j], eachtable.vrPoints[i], eachtable.hrPoints[j+1], eachtable.vrPoints[i+1]]        				
                                    cols.append(cell)
                                    cv2.rectangle(tablecells, (cell[0], cell[1]), (cell[2], cell[3]), (0,0,255), 2)
                    rowCells.append(cols)
            Tables.append(rowCells)

    _exportData2Excel = False
    if(_exportData2Excel):
        #pytesseract
        import tesseract_Test   
        ''' tesseract ocr results '''
        #tableOCR = tesseract_Test.tableOCR(tableOCRImg,Tables)
        tableOCR = tesseract_Test.table_to_ocr(tableOCRImg,Tables)
        ''' google ocr results '''
        saveXlstable = resultsSavePath[:-4]
        excelsavefilepath = saveXlstable + "_TableData.xlsx"
        
        if len(tableOCR) >= 1 :
            savetable = resultsSavePath[:-4]
            savetable = savetable + "_TableOutput.jpg"            
            cv2.imwrite(savetable, finalTableImg)
            print("excelpath:",excelsavefilepath)
            writeToExcel_ColFixed(excelsavefilepath,tableOCR)
        #writeToExcel(excelsavefilepath,tableOCR)

    return isCellBlobs, finalTableImg,Tables,ImageTables

def isIntersect(r1, r2, pointr2):
    res =  r1.overlaps_with(r2)
    if res:
        respoint = r1.is_point_inside_rect(pointr2)
        return  respoint
    else:
        return res

def popSingleBlobsInsideTable(tablelines, singleblobs):
    tableroiLines = []
    
    for i,eachblob in enumerate(tablelines): 
        rowx = eachblob[0];  rowy = eachblob[1]
        rowW = eachblob[2] 
        rowH = eachblob[3]
        r1 = Rect(rowx, rowy, rowW ,rowH)
#        print()
#        print("tablelines{}",eachblob)
        #tableroiLines.append([rowx, rowy, rowW, rowH])
        for i, vrblob in enumerate(singleblobs):
            vx = vrblob[0]; vy = vrblob[1]
            vw = vrblob[2] - vrblob[0] ; vh = vrblob[3] - vrblob[1]
            r2 = Rect(vx, vy, vw, vh)
            midpointx = int(vw/2)
            midpointy = int(vh/2)
            pointr2 = Point(vx+midpointx ,vy+midpointy)
            res =  r1.overlaps_with(r2)
            if res:
                respoint = r1.is_point_inside_rect(pointr2)
                if respoint:
#                    print("single blobs", vrblob)
                    tableroiLines.append([vx, vy, vrblob[2], vrblob[3]])

    tableroiLines.sort(key=lambda x: x[1])    
    return tableroiLines

def drawImgFromBlobs(wordBlobs, imgh, imgw):
    blobsDrawn = np.zeros((imgh,imgw), np.uint8)
    for i,ca in enumerate(wordBlobs):       
        cv2.rectangle(blobsDrawn, (ca[0], ca[1]), (ca[2], ca[3]), 255, cv2.FILLED) 
    return blobsDrawn

    
def getWhiteSpace_Roi(outPath,denoisedImg):
    #imgpath = r'C:\ICDAR\Code\WhiteSpace\whitespaceImgs\cTDaR_t10222_whiteSpaceHr.jpg'
    head, tail = os.path.split(outPath)
    filename =  os.path.splitext(tail)[0]
    imgh, imgw = denoisedImg.shape[:2]
#    imgpath = r'C:\ICDAR\from_Bhasker\whitespacecode\Test22\whitespaceImg'
#    imgpath = r'C:\ICDAR\from_Bhasker\whitespacecode\Test22\whitespaceImg'
#    imgpath =  os.path.join(imgpath, filename + "_whiteSpace_drawn.jpg")
#    whiteSpaceImg = cv2.imread(imgpath)
    print("white space in")
    whiteSpaceImg = whiteSpace.process(denoisedImg, outPath)
    print("white space out")    
    whiteSpaceImg = preprocess(whiteSpaceImg)
    wordSpaceImage = cv2.cvtColor(whiteSpaceImg, cv2.COLOR_BGR2GRAY)
    
    wordblobs = morphCloseBlobs(wordSpaceImage)  
    wordblobs.sort(key=lambda x: x)
    if isDebugImg:
        save_stats = head +"\\"+ filename + "_WSpace_blobs.png"
        cv2.imwrite(save_stats, wordSpaceImage)
    '''Horizontal projection'''
    blobs_Proj_Img = horizontalProjection(wordSpaceImage, wordblobs)

    '''Table rows identification'''
    tablerows = candidateTableRows(blobs_Proj_Img, wordblobs)

    tablelines, srclineblob, singleblobs = findTableRows(tablerows, denoisedImg.copy())
    
    if isDebugImg:
        save_stats = head +"\\"+ filename + "_WSpace_SB.png"
        cv2.imwrite(save_stats, srclineblob)

    if not tablelines:
        return []
    
    '''vertical maximum threshold'''
    textlines = morphCloseBlobs(blobs_Proj_Img)
    textlines.sort(key=lambda x: x[1])
    print("white space")
#    gmax = getGmax(tablelines)
    gmax = 25#25
#    sinleblobs1 = singleblobs.copy()
#    for i, blob in enumerate(singleblobs):
#        x, y, r, b = stats2rect(blob)
#        if (r-x) > imgw*0.4:
#            sinleblobs1.remove(blob)
#    sinleblobs1 = singleblobs.copy()
#    rowlines = rowCorrection(tablelines, sinleblobs1, gmax)
    rowlines = tablelines # skip single white space blobs
    rowlines.sort(key=lambda x: x[1])    
    print('gmax = ',gmax)

    table_blank_image = np.zeros((imgh,imgw), np.uint8)
    for i,ca in enumerate(rowlines):       
        cv2.rectangle(table_blank_image, (ca[0], ca[1]), (ca[2], ca[3]), 255, cv2.FILLED)    
    
    TableRowblobs = morphCloseBlobs(table_blank_image) 
    TableRowblobs.sort(key=lambda x: x)

    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, int(gmax)))
    multitableimg = cv2.morphologyEx(table_blank_image, cv2.MORPH_CLOSE, kernel)
    if isDebugImg:
        save_stats = head +"\\"+ filename + "_WSpace_TableRowBlobs.png"
        cv2.imwrite(save_stats, multitableimg)
    tableblobs = []
    contours, hierarchy = cv2.findContours(multitableimg,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_NONE)
    
    wUhalf = int(imgw * 0.55)
    wLhalf = int(imgw * 0.45)
    rUalign = int(imgw * 0.55)
    rLalign = int(imgw * 0.40)
    for contour in contours:
            x, y, w, h = cv2.boundingRect(contour) 
            roiRect = []
            if x < int(imgw/2) and (x+w) < wUhalf and (x+w) > wLhalf:
                roiRect = [0, y, (x+w)-20, h] 
            elif (x) > rLalign:
                roiRect = [int(imgw/2), y, (x+w), h]
            else:
                roiRect = [0, y, imgw, h]
            tableblobs.append(roiRect) 
    return tableblobs

def filterWordBlobs(word_Blob_Image, whiteSpaceRoi):
    imgh, imgw = word_Blob_Image.shape[:2]
    wordblobs = morphCloseBlobs(word_Blob_Image)
    word_Blob_emptyImg = np.zeros((imgh,imgw), np.uint8)
    
    for i, ca in enumerate(whiteSpaceRoi):
        r1 = Rect(ca[0], ca[1], ca[2], ca[3])
        for j, cb in enumerate(wordblobs):
               r2 = Rect(cb[0], cb[1], cb[2], cb[3])
               res =  r1.overlaps_with(r2)
               if res:
                   cv2.rectangle(word_Blob_emptyImg, (cb[0], cb[1]), (cb[0] + cb[2], cb[1] + cb[3]), 255, cv2.FILLED)
#    cv2.imwrite("word_Blob_emptyImg.jpg", word_Blob_emptyImg)
    return word_Blob_emptyImg
    
def processtable(denoisedImg, outPath, inputSrcImg):
#    print(denoisedImg.shape)
#    denoisedImg = ~denoisedImg
    
    imgh, imgw = denoisedImg.shape[:2]
    head, tail = os.path.split(outPath)
    filename =  os.path.splitext(tail)[0] 
    '''preprocessing'''
    preProcessedImg = preprocess(denoisedImg)
    grayImage = cv2.cvtColor(preProcessedImg, cv2.COLOR_BGR2GRAY)
#    threshImg = cv2.threshold(grayImage, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)[1]
#    threshImg = borderNoiseRemoval(threshImg, 50)
    '''line removal'''
#    imgpath = r'C:\ICDAR\Code\Updated\TRACKB2\cTDaR_t10093.jpg'
#    linesrc = cv2.imread(imgpath)
#    linesrc = preprocess(linesrc)
#    linesrc = cv2.cvtColor(linesrc, cv2.COLOR_BGR2GRAY)
    threshImg = cv2.threshold(grayImage, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)[1]
    threshImg = borderNoiseRemoval(threshImg, 50)
#    linesRemovedImg = lineRemoval(threshImg)
    
    threshImg = ~threshImg
    linesRemovedImg,linesImg = Imaging.Line_Removal_fn(threshImg)    
    linesRemovedImg = ~linesRemovedImg
    if 1==1:
        print("lines removed img processed.")
        save_stats = head +"\\"+ filename + "_linesImg.png"
        print(save_stats)
        cv2.imwrite(save_stats, linesImg) 
        
    #cv2.imwrite(r'C:\ICDAR\Code\Updated\TRACKB2\cTDaR_t10093_lremoved.jpg', linesRemovedImg)
#    linesRemovedImg1 = lineRemoval(threshImg)
#    cv2.imwrite(r'C:\ICDAR\Code\Updated\TRACKB2\cTDaR_t10093_lremoved.jpg', linesRemovedImg1)
    '''connected component analysis'''
    cc_stats = connectedComp_Analysis(linesRemovedImg)
    if len(cc_stats) < 4:  
        return None, [[]]
    del cc_stats[0]   #deleting denoisedImg as blob
    '''histogram based hrsize and second-hump'''
    hrsize, secondHump, src,ccBB = getWordsGap(cc_stats, denoisedImg.copy())
    

    if isDebugImg:
        save_stats = head +"\\"+ filename + "_ccstats_SA.png"
        cv2.imwrite(save_stats, src) 
        save_stats = head +"\\"+ filename + "_ccBB.png"
        cv2.imwrite(save_stats, ~ccBB)
        
#    hrsize = 10
    print("hrsize: {} and second hump:{}".format(hrsize, secondHump))
    
    ''' White Space Roi '''
    
    whiteSpaceRoi = getWhiteSpace_Roi(outPath,denoisedImg)

    print("white space roi: ".format(whiteSpaceRoi))
    
    '''text line formation'''
    word_Blob_Image, secondHumpwordblobs = textLines(hrsize, secondHump, linesRemovedImg, outPath)
    
    ''' Filter word blobs from white space ROI '''
#    whiteSpaceRoi = []
    print("whiteSpaceRoi:".format(len(whiteSpaceRoi)))

    if len(whiteSpaceRoi) > 0:
        print(whiteSpaceRoi)
#        print("White space ROI.")
        word_Blob_Image = filterWordBlobs(word_Blob_Image, whiteSpaceRoi)
        
    if isDebugImg:
        save_stats = head +"\\"+ filename + "_word_Blob_Image.png"
        cv2.imwrite(save_stats, word_Blob_Image)
        
    wordblobs = morphCloseBlobs(word_Blob_Image)  
    wordblobs.sort(key=lambda x: x)
    if isDebugImg:
        save_stats = head +"\\"+ filename + "_singleblobs_SB.png"
        cv2.imwrite(save_stats, word_Blob_Image)
    '''Horizontal projection'''
    blobs_Proj_Img = horizontalProjection(word_Blob_Image, wordblobs)

    '''Table rows identification'''
    tablerows = candidateTableRows(blobs_Proj_Img, wordblobs)

    tablelines, srclineblob, singleblobs = findTableRows(tablerows, denoisedImg.copy())
    
    if isDebugImg:
        save_stats = head +"\\"+ filename + "_tablerows_SB.png"
        cv2.imwrite(save_stats, srclineblob)

    if not tablelines:
        return None, [[]]
    
    '''vertical maximum threshold'''
    textlines = morphCloseBlobs(blobs_Proj_Img)
    textlines.sort(key=lambda x: x[1])
#    linegap = []
#    length = len(tablelines)
#    for i in range(0, length-1):
#        x, y, w, h = stats2rect(textlines[i])
#        bottom_TELa = textlines[i][1] + textlines[i][3]
#        top_TELa1 = textlines[i+1][1]
#        gap = round((top_TELa1 - bottom_TELa))
#        linegap.append(gap)
    
    gmax = getGmax(tablelines)
    gmax = 100#30
    sinleblobs1 = singleblobs.copy()
    for i, blob in enumerate(singleblobs):
        x, y, r, b = stats2rect(blob)
        if (r-x) > imgw*0.4:
            sinleblobs1.remove(blob)
#    print('tablelines', len(tablelines))
    rowlines = rowCorrection(tablelines, sinleblobs1, gmax)
    rowlines.sort(key=lambda x: x[1])
#    print('rowlines', len(rowlines))
#    print(rowlines)
#    print('gmax = ',gmax)

    table_blank_image = np.zeros((imgh,imgw), np.uint8)
    for i,ca in enumerate(rowlines):       
        cv2.rectangle(table_blank_image, (ca[0], ca[1]), (ca[2], ca[3]), 255, cv2.FILLED)    
    
    TableRowblobs = morphCloseBlobs(table_blank_image) 
    TableRowblobs.sort(key=lambda x: x)

    if isDebugImg:
        save_stats = head +"\\"+ filename + "_TableRowimg.png"
        cv2.imwrite(save_stats, table_blank_image)
    
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, int(gmax)))
    multitableimg = cv2.morphologyEx(table_blank_image, cv2.MORPH_CLOSE, kernel) #,  iterations=2)
    if isDebugImg:
        save_stats = head +"\\"+ filename + "_TableRowBlobs.png"
        cv2.imwrite(save_stats, multitableimg)
    tableblobs = []
    contours, hierarchy = cv2.findContours(multitableimg,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_NONE)
    for contour in contours:
            x, y, w, h = cv2.boundingRect(contour) 
            tableblobs.append([x, y, w, h])
    
    print("tableblobs: ",tableblobs)
#    if isDebugImg:
#        save_stats = head +"\\"+ filename + "_TableImage.png"
#        cv2.imwrite(save_stats, table)  
    ''' Pop single blobs inside table '''
#    singleBlobList = popSingleBlobsInsideTable(tableblobs, singleblobs)
#    tablelines = tablelines + singleBlobList
#    tablelines.sort(key=lambda x: x[1])  
    
    ''' table row '''
    tables, FinalTables = tableCols(tableblobs,rowlines)
    ''' table col '''
    blank_image = np.zeros((imgh,imgw), np.uint8)
    cols_image = np.zeros((imgh,imgw), np.uint8)
    blank_image, table = findTableCols(FinalTables, inputSrcImg.copy(), secondHumpwordblobs, cols_image, blank_image)

#    if isDebugImg:
#        save_stats = head +"\\"+ filename + "_tablecolsep.png"
#        cv2.imwrite(save_stats, blank_image)
    cellblobs_image = np.zeros((imgh,imgw), np.uint8)
    resultsSavePath = head +"\\"+ filename + ".png"
#    save_stats = head +"\\"+ filename + "_Tablecells.png"
    isCellBlobs , finalTableImg, tableList, ImageTables = createCellBlobs(FinalTables, cellblobs_image, inputSrcImg, '', resultsSavePath)    
    if True:
        save_stats = head +"\\"+ filename + "_FinalTable.png"
        cv2.imwrite(save_stats, table) 
        save_stats = head +"\\"+ filename + "_Tableout.png"
        cv2.imwrite(save_stats, finalTableImg) 
    return tableList, ImageTables